# -*- coding: utf-8 -*-
# @Author : Mr.sun
# @Time : 2020/2/23 11:36
# @File : logincasa.py


import os
from selenium import webdriver
from openpyxl import load_workbook
from time import sleep
from common.config import conf


class logintest:
    def open(self):
        dataname = os.path.join(conf.dataPath, 'logindata.xlsx')
        book = load_workbook(dataname)
        sheet = book.active
        data = {}
        maxr = sheet.max_row
        maxc = sheet.max_column
        for i in range(1, maxc+1):    #循环1-maxc
            data.setdefault(sheet.cell(1, i).value)   #获取每一列的第一行的值  setdefault方法:添加关键字
        for i in range(2, maxr+1):
            for j in range(1, maxc+1):    #一列一列取值
                data[sheet.cell(1, j).value] = sheet.cell(i, j).value   #第一行 第j列的值  =  第i行  第j列的值  (赋值)
            try:
                self.driver = webdriver.Chrome()
                self.driver.get(data['Url'])
                self.driver.find_element_by_xpath('/html/body/div[2]/div/ul[1]/div/div/a[1]').click()

                if data['UserName']!=None:
                    self.driver.find_element_by_name('accounts').send_keys(data['UserName'])
                if data['PassWord']!=None:
                    self.driver.find_element_by_name('pwd').send_keys(data['PassWord'])
                self.driver.find_element_by_name('pwd').submit()
                sleep(1)
                if data['Expected']!=None:
                    msg = self.driver.find_element_by_xpath('//*[@id="common-prompt"]/p').text
                    if msg == data['Expected']:
                        print('test1')
                        sheet.cell(i, maxc, 'PASS')
                    else:
                        sheet.cell(i, maxc, 'FAIL')
                else:
                    print('测试数据有误')
                self.driver.quit()
            except:
                raise
        book.save(dataname)


logintest().open()
